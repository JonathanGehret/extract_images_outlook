#!/usr/bin/env python3
"""I/O helpers: image listing, renaming/backups, and Excel writes.
This module exposes functions that operate on explicit paths/params so they
can be used from the GUI without carrying state.
"""
import os
import re
import shutil
from datetime import datetime, timedelta
import pandas as pd


def _natural_sort_key(filename: str):
    """
    Generate a sort key for natural (human-friendly) sorting.
    Converts numeric parts to integers so "fotofallen_2025_1.jpeg" comes before
    "fotofallen_2025_10.jpeg" instead of alphabetical sorting.
    
    Examples:
        "fotofallen_2025_1.jpeg"   -> ['fotofallen_', 2025, '_', 1, '.jpeg']
        "fotofallen_2025_100.jpeg" -> ['fotofallen_', 2025, '_', 100, '.jpeg']
    """
    parts = re.split(r'(\d+)', filename)
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def get_image_files(images_folder: str):
    if not images_folder:
        return []
    try:
        files = [f for f in os.listdir(images_folder) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
    except FileNotFoundError:
        return []
    return sorted(files, key=_natural_sort_key)


def refresh_image_list(images_folder: str, old_list: list, current_index: int):
    try:
        all_files = os.listdir(images_folder)
        image_files = [f for f in all_files if f.lower().endswith(('.jpg', '.jpeg', '.png')) and not f.startswith('.') and 'backup' not in f.lower()]
        image_files.sort(key=_natural_sort_key)

        old_current_file = old_list[current_index] if current_index < len(old_list) else None
        if old_current_file and old_current_file in image_files:
            current_index = image_files.index(old_current_file)
        elif current_index >= len(image_files):
            current_index = max(0, len(image_files) - 1)

        return image_files, current_index
    except Exception:
        return old_list, current_index


def create_backup_and_rename_image(images_folder: str, image_file: str, location: str, date,
                                   species1: str, count1: str,
                                   species2: str, count2: str,
                                   species3: str, count3: str,
                                   species4: str, count4: str,
                                   new_id: int,
                                   generl_checked: bool = False,
                                   luisa_checked: bool = False,
                                   unspecified_bartgeier: bool = False):
    try:
        old_path = os.path.join(images_folder, image_file)
        if not os.path.exists(old_path):
            return None

        backup_dir = os.path.join(images_folder, "backup_originals")
        os.makedirs(backup_dir, exist_ok=True)
        backup_path = os.path.join(backup_dir, image_file)
        if not os.path.exists(backup_path):
            shutil.copy2(old_path, backup_path)

        animals = []
        for species, count in [
            (species1, count1),
            (species2, count2),
            (species3, count3),
            (species4, count4)
        ]:
            normalized = str(species).strip() if species is not None else ""
            if not normalized:
                continue

            if "bartgeier" in normalized.lower():
                continue

            count_suffix = ""
            if count not in (None, ""):
                try:
                    count_val = int(float(count))
                    if count_val > 1:
                        count_suffix = f"_{count_val}"
                except (ValueError, TypeError):
                    pass

            lower_name = normalized.lower()
            if normalized.upper() == 'RK':
                animals.append(f"Rabenkrähe{count_suffix}")
            elif lower_name == 'rk':
                animals.append(f"Kolkrabe{count_suffix}")
            elif normalized.upper() == 'RV':
                animals.append(f"Kolkrabe{count_suffix}")
            elif lower_name in ['fuchs', 'marder']:
                animals.append(f"{normalized.capitalize()}{count_suffix}")
            elif lower_name == 'gams':
                animals.append(f"Gämse{count_suffix}")
            else:
                animals.append(f"{normalized}{count_suffix}")

        if unspecified_bartgeier:
            animals.append("Unbestimmt_Bartgeier")

        animal_str = "_".join(animals)

        try:
            # Normalize date to string in DD.MM.YYYY first
            if isinstance(date, datetime):
                normalized = date
            elif isinstance(date, (int, float)):
                normalized = datetime(1899, 12, 30) + timedelta(days=float(date))
            else:
                date_str_raw = str(date).strip()
                if date_str_raw.replace('.', '').isdigit() and '.' not in date_str_raw:
                    # Excel serial stored as plain integer string
                    normalized = datetime(1899, 12, 30) + timedelta(days=float(date_str_raw))
                elif '.' in date_str_raw:
                    parts = date_str_raw.split('.')
                    if len(parts) == 3:
                        day, month, year = parts
                        if len(year) == 2:
                            year = '20' + year
                        normalized = datetime(int(year), int(month), int(day))
                    else:
                        normalized = datetime.fromisoformat(date_str_raw)
                else:
                    normalized = datetime.fromisoformat(date_str_raw)
            day = normalized.day
            month = normalized.month
            year = normalized.year % 100
            date_str = f"{month:02d}.{day:02d}.{year:02d}"
        except Exception:
            date_str = str(date)

        if not date_str:
            date_str = "unknown-date"

        special_names = []
        if generl_checked:
            special_names.append("Generl")
        if luisa_checked:
            special_names.append("Luisa")

        special_str = "_".join(special_names)

        parts = [
            location,
            f"{int(new_id):04d}",
            date_str
        ]

        if special_str:
            parts.append(special_str)
        if animal_str:
            parts.append(animal_str)

        new_name = "_".join([p for p in parts if p]) + ".jpeg"

        new_path = os.path.join(images_folder, new_name)

        counter = 1
        base_root, base_ext = os.path.splitext(new_name)
        if not base_ext:
            base_ext = ".jpeg"
        while os.path.exists(new_path) and new_path != old_path:
            new_name = f"{base_root}_{counter}{base_ext}"
            new_path = os.path.join(images_folder, new_name)
            counter += 1

        if old_path != new_path:
            os.rename(old_path, new_path)
        return new_name
    except Exception:
        return None


def get_next_id_for_location(output_excel: str, location: str):
    try:
        df = pd.read_excel(output_excel, sheet_name=location)
        if not df.empty and 'Nr. ' in df.columns:
            numeric_ids = []
            for id_val in df['Nr. ']:
                try:
                    if pd.notna(id_val):
                        numeric_ids.append(int(float(id_val)))
                except (ValueError, TypeError):
                    continue
            if numeric_ids:
                return max(numeric_ids) + 1
            return 1
        return 1
    except Exception:
        return 1


def save_single_result(excel_path, location, data):
    """Save a single result to Excel while preserving ALL original formatting."""
    import openpyxl
    
    try:
        # Open existing workbook or create new one
        try:
            workbook = openpyxl.load_workbook(excel_path)
        except FileNotFoundError:
            # Create new workbook with standard structure
            workbook = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Create sheets for each location
            for loc in ['FP1', 'FP2', 'FP3', 'Nische']:
                ws = workbook.create_sheet(loc)
                # Add standard headers
                headers = ['Nr. ', 'Datum', 'Uhrzeit', 'Generl', 'Luisa', 'Unbestimmt',
                          'Aktivität', 'Art 1', 'Anz. 1', 'Art 2', 'Anz. 2',
                          'Art 3', 'Anz. 3', 'Art 4', 'Anz. 4', 'Interaktion', 'Sonstiges']
                ws.append(headers)
        
        # Get or create the worksheet for this location
        if location not in workbook.sheetnames:
            ws = workbook.create_sheet(location)
            # Add headers if new sheet
            headers = ['Nr. ', 'Datum', 'Uhrzeit', 'Generl', 'Luisa', 'Unbestimmt',
                      'Aktivität', 'Art 1', 'Anz. 1', 'Art 2', 'Anz. 2',
                      'Art 3', 'Anz. 3', 'Art 4', 'Anz. 4', 'Interaktion', 'Sonstiges']
            ws.append(headers)
        else:
            ws = workbook[location]
        
        # Find the actual last row with data (not just max_row which can be misleading)
        actual_last_row = 1  # Start with header row
        for row_num in range(2, ws.max_row + 1):  # Start from row 2 (after headers)
            # Check if any cell in this row has actual data
            has_data = False
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            
            if has_data:
                actual_last_row = row_num
            # If we find several empty rows in a row, we can break early
            elif row_num > actual_last_row + 10:  # Allow some gap but not too much
                break
        
        # Next row is right after the actual last row with data
        next_row = actual_last_row + 1
        
        # Read existing headers from row 1
        headers = []
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col)
            headers.append(header_cell.value or '')
        
        # Create column mapping (from data keys to Excel headers)
        column_mapping = {
            'Nr. ': 'Nr. ',
            'Datum': 'Datum', 
            'Uhrzeit': 'Uhrzeit',
            'Generl': 'Generl',
            'Luisa': 'Luisa', 
            'Unbestimmt': 'Unbestimmt',
            'Aktivität': 'Aktivität',
            'Art 1': 'Art 1',
            'Anzahl 1': 'Anz. 1',  # Map data key to Excel header
            'Art 2': 'Art 2',
            'Anzahl 2': 'Anz. 2',  # Map data key to Excel header
            'Art 3': 'Art 3',      # New
            'Anzahl 3': 'Anz. 3',  # New
            'Art 4': 'Art 4',      # New
            'Anzahl 4': 'Anz. 4',  # New
            'Interaktion': 'Interaktion',
            'Sonstiges': 'Sonstiges'
        }
        
        # Write data to the new row
        for col_idx, header in enumerate(headers, 1):
            # Find matching data key for this Excel header
            data_key = None
            for data_k, excel_h in column_mapping.items():
                if excel_h == header:
                    data_key = data_k
                    break
            
            # Write the data value if we have a match
            if data_key and data_key in data:
                cell = ws.cell(row=next_row, column=col_idx)
                
                # Copy formatting from the row above first (if it exists)
                if actual_last_row > 1:  # Make sure there's a previous data row
                    source_cell = ws.cell(row=actual_last_row, column=col_idx)
                    
                    # Copy cell formatting
                    if source_cell.has_style:
                        cell.font = openpyxl.styles.Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color
                        )
                        cell.fill = openpyxl.styles.PatternFill(
                            fill_type=source_cell.fill.fill_type,
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color
                        )
                        cell.border = openpyxl.styles.Border(
                            left=source_cell.border.left,
                            right=source_cell.border.right,
                            top=source_cell.border.top,
                            bottom=source_cell.border.bottom
                        )
                        cell.alignment = openpyxl.styles.Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical
                        )
                        cell.number_format = source_cell.number_format
                
                # Set the value - simple and direct
                if data_key in ['Generl', 'Luisa'] and data[data_key] == 'X':
                    cell.value = 'x'  # Use lowercase x instead of uppercase X
                else:
                    cell.value = data[data_key]  # Write exactly as received
        
        # Save the workbook (preserves ALL original formatting)
        workbook.save(excel_path)
        print(f"✅ Data saved to {excel_path} (sheet: {location}) at row {next_row} - formatting preserved")
        return True
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        import traceback
        traceback.print_exc()
        raise
